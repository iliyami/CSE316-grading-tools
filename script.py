from datetime import datetime, timedelta
import os
import sys
import pandas as pd
import shutil
import re
import subprocess
from openpyxl import load_workbook

DEFAULT_DEADLINE_SHIFT = 0
LATE_SUBMISSIONS_FILE = "late_submissions.txt" 


def display_intro():
    """Displays a beautifully formatted introduction."""
    intro_text = """
    ░▒▓██████▓▒░ ░▒▓███████▓▒░▒▓████████▓▒░      ░▒▓███████▓▒░   ░▒▓█▓▒░░▒▓███████▓▒░ 
    ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░      ░▒▓█▓▒░                    ░▒▓█▓▒░▒▓████▓▒░▒▓█▓▒░        
    ░▒▓█▓▒░      ░▒▓█▓▒░      ░▒▓█▓▒░                    ░▒▓█▓▒░  ░▒▓█▓▒░▒▓█▓▒░        
    ░▒▓█▓▒░       ░▒▓██████▓▒░░▒▓██████▓▒░        ░▒▓███████▓▒░   ░▒▓█▓▒░▒▓███████▓▒░  
    ░▒▓█▓▒░             ░▒▓█▓▒░▒▓█▓▒░                    ░▒▓█▓▒░  ░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ 
    ░▒▓█▓▒░░▒▓█▓▒░      ░▒▓█▓▒░▒▓█▓▒░                    ░▒▓█▓▒░  ░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ 
    ░▒▓██████▓▒░░▒▓███████▓▒░░▒▓████████▓▒░      ░▒▓███████▓▒░   ░▒▓█▓▒░░▒▓██████▓▒░  
                                                                                   
                            Grading Setup Automation
                            Created by: Iliya Mirzaei
    -------------------------------------------------------------------------------------
    """
    print(intro_text)

def extract_name(email):
    """Extracts first and last name from the email format x_y.z@q.d"""
    match = re.match(r".*_(\w+)\.(\w+)@", email)
    if match:
        return match.group(1).capitalize(), match.group(2).capitalize()
    return None, None

def update_excel_file(file_path, mapping):
    """Updates the Excel file without changing formatting, placing values in the next column."""
    wb = load_workbook(file_path)
    ws = wb.active  # Assuming a single-sheet Excel file

    for row in ws.iter_rows():
        for cell in row:
            if cell.value in mapping:
                next_col_index = cell.column + 1  # Move to the next column
                ws.cell(row=cell.row, column=next_col_index, value=mapping[cell.value])

    wb.save(file_path)

def read_extensions(extension_file):
    """Reads the extension file and returns a dictionary {group_name: new_deadline}."""
    extensions = {}
    if extension_file and os.path.exists(extension_file):
        with open(extension_file, "r") as file:
            for line in file:
                match = re.match(r".* - (.*), (\d{2}/\d{2}/\d{4})", line.strip())
                if match:
                    group_name, deadline = match.groups()
                    extensions[group_name] = datetime.strptime(deadline, "%m/%d/%Y")
    return extensions


def get_latest_commit_timestamp(repo_path):
    """Returns the timestamp of the latest commit in the cloned repo."""
    try:
        timestamp = subprocess.check_output(
            ["git", "log", "-1", "--format=%ct"], cwd=repo_path
        ).decode("utf-8").strip()
        return datetime.fromtimestamp(int(timestamp))
    except Exception as e:
        print(f"⚠️ Error retrieving commit timestamp for {repo_path}: {e}")
        return None


def calculate_penalty(submission_time, deadline, grace_period):
    """Calculates the penalty based on the submission time and deadline."""
    time_diff = (submission_time - deadline).total_seconds() / 3600  # Convert to hours

    if time_diff <= grace_period:
        return "✅ On time", None, time_diff
    elif 6 < time_diff <= 30:
        return "⏳ Late (90% cap)", 90, time_diff
    elif 30 < time_diff <= 54:
        return "⏳ Late (80% cap)", 80, time_diff
    elif 54 < time_diff <= 78:
        return "⏳ Late (70% cap)", 70, time_diff
    elif 78 < time_diff <= 102:
        return "⏳ Late (60% cap)", 60, time_diff
    else:
        return "⏳ Late (50% cap)", 50, time_diff
    
def log_late_submission(log_data):
    """Logs late submission details into a file."""
    with open(LATE_SUBMISSIONS_FILE, "a") as file:
        file.write(log_data + "\n")

def process_grading(grader_name, test_file_path, team_grading_path, deadline_str, grace_period, extension_file):
    display_intro()

    # Convert deadline to datetime object
    deadline = datetime.strptime(deadline_str, "%m-%d-%Y")
    extensions = read_extensions(extension_file) if extension_file else {}

    # Clear previous late submissions log
    with open(LATE_SUBMISSIONS_FILE, "w") as file:
        file.write("Late Submissions Log\n--------------------\n")

    # Read the team grading file which includes details about graders and teams
    df = pd.read_excel(team_grading_path)

    # Filter rows only for the given grader name - Feel free to remove it if you want to help other TAs :D
    filtered_df = df[df["Grader Name"] == grader_name]

    # Get unique groups
    unique_groups = filtered_df["group_name"].unique()

    teams_path = os.path.join(os.getcwd(), 'teams')
    if os.path.exists(teams_path):
        shutil.rmtree(teams_path, ignore_errors=True)
    os.makedirs(teams_path, exist_ok=False)

    for group in unique_groups:
        # Get relevant group members
        group_members = filtered_df[filtered_df["group_name"] == group] 

        # Extract repo URL and convert HTTP to SSH
        repo_url_http = group_members["student_repository_url"].iloc[0]
        repo_url_ssh = repo_url_http.replace("https://", "git@").replace("/", ":", 1)

        # Define the directory name for the cloned repo
        repo_name = repo_url_http.split("/")[-1].replace(".git", "")
        repo_path = os.path.join(teams_path, repo_name)

        # Clone the repository using SSH
        if not os.path.exists(repo_path):  # Avoid cloning if the repo already exists
            subprocess.run(["git", "clone", repo_url_ssh], cwd=teams_path)

        # Ensure the repo was cloned successfully
        if not os.path.exists(repo_path):
            print(f"❌ Failed to clone repository: {repo_url_ssh}")
            continue

        # Copy the test file to the cloned repo
        dest_test_file = os.path.join(repo_path, os.path.basename(test_file_path))
        shutil.copy(test_file_path, dest_test_file)

        # Extract student names
        students = []
        for email in group_members["roster_identifier"]:
            first_name, last_name = extract_name(email)
            if first_name and last_name:
                students.append(f"{first_name} {last_name}")

        # Get grader email
        grader_email = group_members["Grader Email"].iloc[0]

        # Fill the test file
        test_df = pd.read_excel(dest_test_file, header=None)
        mapping = {
            "Team Name:": group,
            "Student 1:": students[0] if len(students) > 0 else "?",
            "Student 2:": students[1] if len(students) > 1 else "?",
            "Grader Name:": grader_name,
            "Grader Email:": grader_email
        }

        update_excel_file(dest_test_file, mapping)

        submission_time = get_latest_commit_timestamp(repo_path)
        if submission_time:
            team_deadline = extensions.get(group, deadline)
            status, score_cap, diff_hours = calculate_penalty(submission_time, team_deadline, grace_period)

            log_data = (
                f"Team: {group}\n"
                f"   - 🕒 Submission Timestamp: {submission_time}\n"
                f"   - 📝 Original Deadline: {deadline}\n"
                f"   - ⏳ Time Difference: {diff_hours:.2f} hours\n"
                f"   - 🚦 Submission Status: {status}"
            )

            if group in extensions:
                log_data += f"\n   - 🆓 Extension Granted Until: {team_deadline}"

            if score_cap:
                log_data += f"\n   - 🔴 Max Score Capped at {score_cap}%"

            log_late_submission(log_data)

            print(log_data + "\n")

        print(f"Successfully updated grading file in {repo_name}")
        print("-----------------------------------------------\n")


if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Usage: python script.py <Grader Name> <test_file.xlsx> <team_grading.xlsx> <Deadline MM-DD-YYYY> [Shift (default: 0 hours - no shift)] [extension_file.txt]")
        sys.exit(1)

    grader_name = sys.argv[1]
    test_file_path = sys.argv[2]
    team_grading_path = sys.argv[3]
    deadline_str = sys.argv[4]
    grace_period = int(sys.argv[5]) if len(sys.argv) > 5 else DEFAULT_DEADLINE_SHIFT
    extension_file = sys.argv[6] if len(sys.argv) > 6 else None

    process_grading(grader_name, test_file_path, team_grading_path, deadline_str, grace_period, extension_file)
